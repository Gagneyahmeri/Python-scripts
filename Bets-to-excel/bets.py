import pandas as pd
from bs4 import BeautifulSoup, Tag
import json
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.formatting.rule import CellIsRule, FormulaRule


unitSize = 80  # Define your unit size here

tonybet_file_path = "C:/Your/Path/Here" # Define file path for tonybet data
betinia_file_path = "C:/Your/Path/Here" # Define file path for betinia data

excel_file_path = "C:/You/Path/Here/excel.xlxs" #Define your excel where bets will be saved

"""Tonybet parsing"""

def parseTeams(tonybet_file_path):
    with open(tonybet_file_path, encoding='utf-8') as fp:
        soup = BeautifulSoup(fp, 'html.parser')

    class_names = "BetDetails-module_teams__Thy6v column-common-module_bold700__s-QDx"

    # Find all div elements with the specified class names
    elements = soup.find_all("div", class_=class_names)

    team_names = []

    for element in elements:
        if element.find('br'):
            contents = element.contents
            names = []
            for content in contents:
                if not isinstance(content, Tag):
                    names.extend(content.split('<br/>'))
                else:
                    pass
            # Concatenate the team names with ' vs ' and add to the team_names list
            if len(names) == 2:  # Ensure there are exactly two names to concatenate
                match_str = f'{names[0].strip()} vs {names[1].strip()}'
                team_names.append(match_str)

    return team_names


def parseBetStakes(tonybet_file_path, unitSize):
    with open(tonybet_file_path, encoding='utf-8') as fp:
        soup = BeautifulSoup(fp, 'html.parser')

    # Find all span elements with the specified data-test attribute
    stakes_elements = soup.find_all("span", attrs={"data-test": "bet-stake-fullAmount"})

    stakes = []

    for element in stakes_elements:
        stake_value = element.text.strip()
        normalized_stake = float(stake_value) / unitSize
        stakes.append(normalized_stake)

    return stakes


def parseWinLossValues(tonybet_file_path, unitSize):
    with open(tonybet_file_path, encoding='utf-8') as fp:
        soup = BeautifulSoup(fp, 'html.parser')

    # Selector for the <td> elements that may contain wins
    td_selector = 'td.cabinet-history-module_td__feqKW[data-test="betting-history-winFull"]'
    
    # Find all <td> elements that match the selector
    td_elements = soup.select(td_selector)

    win_loss_units = []

    for td in td_elements:
        win_span = td.find("span", class_="BetStake-module_win__bq281")
        
        if win_span:
            win_amount = float(win_span.text.strip().replace(',', '.'))
            win_units = win_amount / unitSize
        else:
            win_units = 0
        
        win_loss_units.append(win_units)

    return win_loss_units


def parseMatchOdds(tonybet_file_path):
    with open(tonybet_file_path, encoding='utf-8') as fp:
        soup = BeautifulSoup(fp, 'html.parser')

    # Selector for the <td> elements containing odds
    odds_selector = 'td.cabinet-history-module_td__feqKW[data-test="betting-history-odds"]'
    
    # Find all <td> elements that match the selector
    odds_elements = soup.select(odds_selector)

    match_odds = []

    for td in odds_elements:
        odds_value = float(td.text.strip())
        match_odds.append(odds_value)

    return match_odds

""" Tonybet parsing ends"""

def createNewSheet(excel_file_path, base_sheet_name="Week", column_widths=None):
    wb = load_workbook(excel_file_path)
    existing_sheet_names = wb.sheetnames
    week_sheets = [sheet_name for sheet_name in existing_sheet_names if base_sheet_name in sheet_name]

    last_week = max([int(sheet_name.replace(base_sheet_name, '').strip()) for sheet_name in week_sheets], default=0)
    new_sheet_name = f"{base_sheet_name} {last_week + 1}"
    
    ws = wb.create_sheet(title=new_sheet_name)
    
    if column_widths:
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
    
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center')
    right_alignment = Alignment(horizontal='right')
    
    headers = ['Match', 'Stake', 'Result', 'Odds', 'Return']
    for col, header in zip('ABCDE', headers):
        cell = ws[f'{col}1']
        cell.value = header
        cell.font = bold_font
        if col in 'BCDE':
            cell.alignment = center_alignment
    
    totals_headers = {"G1": "Totals", "G2": "Wagered", "G3": "Returned", "G5": "Profit in units"}
    for cell_ref, header in totals_headers.items():
        cell = ws[cell_ref]
        cell.value = header
        cell.font = bold_font

    stake_formula = "=ROUND(SUM(B2:B500), 2)"
    ws['H2'] = stake_formula
    ws['H2'].alignment = center_alignment

    return_formula = "=ROUND(SUM(E2:E500), 2)"
    ws['H3'] = return_formula
    ws['H3'].alignment = center_alignment

    ws['I5'] = "In €"
    ws['I5'].font = bold_font
    ws['I5'].alignment = right_alignment

    unit_formula = "=ROUND(H3-H2, 2)"
    ws['H5'] = unit_formula
    ws['H5'].alignment = right_alignment

    green_fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
    red_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    ws.conditional_formatting.add('H5', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add('H5', CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=green_fill))

    profit_formula = f"=H5*{unitSize}"
    ws['J5'] = profit_formula
    ws['J5'].alignment = center_alignment
    ws['J5'].number_format = '#,##0.00 €'

    ws.conditional_formatting.add('J5', CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add('J5', CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=green_fill))

    ws['G8'].value = "Average Odds"
    ws['G8'].font = bold_font

    avg_odds_formula = "=ROUND(AVERAGE(D2:D500), 2)"
    ws['H8'] = avg_odds_formula
    ws['H8'].alignment = center_alignment

    wb.save(excel_file_path)


def appendToExcel(df, excel_file_path, sheet_name):
    try:
        wb = load_workbook(excel_file_path)
        existing_sheet_names = wb.sheetnames
        if sheet_name not in existing_sheet_names:
            print(f"Sheet '{sheet_name}' not found. Exiting...")
            return
        ws = wb[sheet_name]
    
        # Find the next available row in the sheet
        next_row = ws.max_row + 1 if ws.max_row > 0 else 1
    
        for _, row in df.iterrows():
            ws.cell(row=next_row, column=1, value=row['A'])
            ws.cell(row=next_row, column=2, value=row['B'])
            ws.cell(row=next_row, column=4, value=row['D'])
            ws.cell(row=next_row, column=5, value=row['E'])
            next_row += 1
    
        wb.save(excel_file_path)
    except FileNotFoundError:
        print("Excel file not found. Exiting...")
        return


def parseTonyBetToExcel(html_file_path, unitSize, excel_file_path):
    try:
        wb = load_workbook(excel_file_path)
    except FileNotFoundError:
        print("Excel file not found. Exiting...")
        return
    
    existing_sheet_names = wb.sheetnames
    
    week_sheets = [sheet_name for sheet_name in existing_sheet_names if "Week" in sheet_name]
    if not week_sheets:
        print("No week sheet found. Please create a new sheet first.")
        return
    
    latest_week = max([int(sheet_name.replace("Week", "").strip()) for sheet_name in week_sheets])
    sheet_name = f"Week {latest_week}"
    dark_green_fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    
    # Get the latest week sheet
    ws = wb[sheet_name]
    
    team_matches = parseTeams(html_file_path)
    bet_stakes = parseBetStakes(html_file_path, unitSize)
    win_loss_units = parseWinLossValues(html_file_path, unitSize)
    match_odds = parseMatchOdds(html_file_path)
    
    combined_data = [{
        "team_names": team_match,
        "bet_stake_units": bet_stake,
        "odds": match_odd,
        "win_loss_units": win_loss_unit,
    } for team_match, bet_stake, win_loss_unit, match_odd in zip(team_matches, bet_stakes, win_loss_units, match_odds)]
    
    # Convert the combined data into a DataFrame
    df_new_data = pd.DataFrame(combined_data)
    
    # Center alignment setup
    center_alignment = Alignment(horizontal='center')
    
    # Find the next available row in the sheet after existing content
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        if row[0].value is None:
            next_row = row[0].row
            break
    else:
        next_row = ws.max_row + 1
    
    df_new_data = df_new_data.iloc[::-1].reset_index(drop=True)

    for _, row in df_new_data.iterrows():
    # Write to column A
        ws.cell(row=next_row, column=1, value=row['team_names'])
    
        # Write to column B and apply center alignment
        cell = ws.cell(row=next_row, column=2, value=row['bet_stake_units'])
        cell.alignment = center_alignment
    
        # Apply the formula to column C and center the cell
        formula = f'=IF(ISBLANK(E{next_row}), "", IF(E{next_row}>0, "W", "L"))'
        cell = ws.cell(row=next_row, column=3, value=formula)
        cell.alignment = center_alignment  # Center alignment is applied here
        ws.conditional_formatting.add(f'C2:C{next_row}',
            CellIsRule(operator='equal', formula=['"W"'], stopIfTrue=True, fill=dark_green_fill))
        ws.conditional_formatting.add(f'C2:C{next_row}',
            CellIsRule(operator='equal', formula=['"L"'], stopIfTrue=True, fill=red_fill))
        
        # Write to column D and apply center alignment
        cell = ws.cell(row=next_row, column=4, value=row['odds'])
        cell.alignment = center_alignment
        
        # Write the actual value to column E and apply center alignment
        cell = ws.cell(row=next_row, column=5, value=row['win_loss_units'])
        cell.alignment = center_alignment
        
        next_row += 1

    wb.save(excel_file_path)
    
    print("Data from TonyBet has been successfully added to the Excel file.")


def parseBetiniaToExcel(betinia_file_path, unitSize, excel_file_path):
    try:
        with open(betinia_file_path, 'r', encoding='utf-8') as json_file:
            betinia_data = json.load(json_file)

        for item in betinia_data:
            if 'Odds' in item:
                # Replace "." with ","
                item['Odds'] = item['Odds'].replace('.', ',')
                # Convert odds to float
                item['Odds'] = float(item['Odds'].replace(',', '.'))

        # Create a new DataFrame to hold the parsed data
        df_new_data = pd.DataFrame(betinia_data)

        try:
            wb = load_workbook(excel_file_path)
        except FileNotFoundError:
            print("Excel file not found. Exiting...")
            return

        existing_sheet_names = wb.sheetnames

        week_sheets = [sheet_name for sheet_name in existing_sheet_names if "Week" in sheet_name]
        if not week_sheets:
            print("No week sheet found. Please create a new sheet first.")
            return

        latest_week = max([int(sheet_name.replace("Week", "").strip()) for sheet_name in week_sheets])
        sheet_name = f"Week {latest_week}"

        # Get the latest week sheet
        ws = wb[sheet_name]

        # Center alignment setup
        center_alignment = Alignment(horizontal='center')
        dark_green_fill = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Find the next available row in the sheet after existing content
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value is None:
                next_row = row[0].row
                break
        else:
            next_row = ws.max_row + 1

        # Reverse the DataFrame order using iloc
        df_new_data = df_new_data.iloc[::-1].reset_index(drop=True)


        for _, row in df_new_data.iterrows():
            # Write to column A
            ws.cell(row=next_row, column=1, value=row['Bets'])

            # Write to column B and apply center alignment
            cell = ws.cell(row=next_row, column=2, value=float(row['StakeAmount'].split()[0]) / unitSize)
            cell.alignment = center_alignment

            # Apply the formula to column C and center the cell
            formula = f'=IF(ISBLANK(E{next_row}), "", IF(E{next_row}>0, "W", "L"))'
            cell = ws.cell(row=next_row, column=3, value=formula)
            cell.alignment = center_alignment
            ws.conditional_formatting.add(f'C2:C{next_row}',
                                           CellIsRule(operator='equal', formula=['"W"'], stopIfTrue=True,
                                                      fill=dark_green_fill))
            ws.conditional_formatting.add(f'C2:C{next_row}',
                                           CellIsRule(operator='equal', formula=['"L"'], stopIfTrue=True,
                                                      fill=red_fill))

            # Write to column D and apply center alignment
            cell = ws.cell(row=next_row, column=4, value=row['Odds'])
            cell.alignment = center_alignment

            # Write the actual value to column E and apply center alignment
            cell = ws.cell(row=next_row, column=5, value=float(row['Return'].split()[0]) / unitSize)
            cell.alignment = center_alignment

            next_row += 1

        wb.save(excel_file_path)

        print("Data from Betinia has been successfully added to the Excel file.")

    except FileNotFoundError:
        print("Betinia JSON file not found. Exiting...")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

def updateTotals(excel_file_path, totals_sheet_name='Totals'):
    wb = openpyxl.load_workbook(excel_file_path)
    

    # Initialize formulas for summing H2, H3, H5, and now J5 across all sheets
    formula_h2 = '='
    formula_h3 = '='
    formula_h5 = '='
    formula_j5 = '='

    for sheet in wb.sheetnames:
        if sheet != totals_sheet_name:
            formula_h2 += f"'{sheet}'!H2+"
            formula_h3 += f"'{sheet}'!H3+"
            formula_h5 += f"'{sheet}'!H5+"
            formula_j5 += f"'{sheet}'!J5+"
    
    # Remove the last plus sign from the formulas
    formula_h2 = formula_h2.rstrip('+')
    formula_h3 = formula_h3.rstrip('+')
    formula_h5 = formula_h5.rstrip('+')
    formula_j5 = formula_j5.rstrip('+')

    # Access the totals sheet
    totals_sheet = wb[totals_sheet_name]

    # Update cells B2 and B3 with the new formulas and center their content
    totals_sheet['B2'].value = formula_h2
    totals_sheet['B2'].alignment = Alignment(horizontal='center')
    totals_sheet['B3'].value = formula_h3
    totals_sheet['B3'].alignment = Alignment(horizontal='center')

    # Update cell B5 with the sum of H5 values from all sheets
    totals_sheet['B5'].value = formula_h5
    totals_sheet['B5'].alignment = Alignment(horizontal='center')

    # Update cell D5 to sum J5 values from all sheets and format as currency
    totals_sheet['D5'].value = formula_j5
    totals_sheet['D5'].number_format = '#,##0.00 €'

    green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')

    # Apply conditional formatting to cell B5
    totals_sheet.conditional_formatting.add('B5', FormulaRule(formula=['B5>0'], stopIfTrue=True, fill=green_fill))
    totals_sheet.conditional_formatting.add('B5', FormulaRule(formula=['B5<0'], stopIfTrue=True, fill=orange_fill))

    # Apply conditional formatting to cell D5
    totals_sheet.conditional_formatting.add('D5', FormulaRule(formula=['D5>0'], stopIfTrue=True, fill=green_fill))
    totals_sheet.conditional_formatting.add('D5', FormulaRule(formula=['D5<0'], stopIfTrue=True, fill=orange_fill))


    wb.save(excel_file_path)
    print(f"Updated totals in sheet '{totals_sheet_name}', formatting cells B2, B3, B5, and summing J5 in D5.")


def main():
    column_widths = {'A': 50, 'B': 11, 'C': 11.5, 'D': 10.5, 'E': 12, 'G': 13}
    while True:
        print("Options:")
        print("1. Add all bets to Excel")
        print("2. Add TonyBet to Excel")
        print("3. Add Betinia to Excel")
        print("4. Update totals")
        print("5. Add new sheet to Excel")
        print("6. Exit")
        
        choice = input("Enter your choice (1/2/3/4/5/6): ").strip()
        
        if choice == '1':
            # Add all bets to Excel
            parseTonyBetToExcel(tonybet_file_path, unitSize, excel_file_path)
            parseBetiniaToExcel(betinia_file_path, unitSize, excel_file_path)
        elif choice == '2':
            # Add TonyBet to Excel
            parseTonyBetToExcel(tonybet_file_path, unitSize, excel_file_path)
        elif choice == '3':
            # Add Betinia to Excel
            parseBetiniaToExcel(betinia_file_path, unitSize, excel_file_path)
        elif choice == '4':
            # Update totals
            updateTotals(excel_file_path)
        elif choice == '5':
            # Add new sheet to Excel
            createNewSheet(excel_file_path, "Week", column_widths)
        elif choice == '6':
            # Exit the program
            print("Exiting...")
            break
        else:
            print("Invalid choice. Please enter a number between 1 and 6.")

if __name__ == "__main__":
    main()
